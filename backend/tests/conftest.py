"""
Felles pytest-fixturer.

Vi kjører hver test mot en frisk database for å unngå at testene
påvirker hverandre. Schema bygges opp via `Base.metadata.create_all`
slik at vi slipper å kjøre Alembic per test.
"""

from __future__ import annotations

import os
from collections.abc import AsyncIterator
from pathlib import Path

import pytest
import pytest_asyncio
from fastapi import FastAPI
from httpx import ASGITransport, AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

# Sett miljøvariabler FØR vi importerer app-en så de plukkes opp av Settings.
#
# Viktig sikkerhetsbarriere:
# Tester skal ALDRI kjøre mot dev/prod-database. Hvis TEST_DATABASE_URL ikke er
# satt eksplisitt, bruker vi en separat *_test-database i samme docker-nettverk.
_DEFAULT_TEST_DATABASE_URL = (
    "postgresql+asyncpg://xlent:xlent_dev_password@postgres:5432/xlent_compliance_test"
)
os.environ["DATABASE_URL"] = os.environ.get(
    "TEST_DATABASE_URL",
    _DEFAULT_TEST_DATABASE_URL,
)
if "_test" not in os.environ["DATABASE_URL"]:
    raise RuntimeError(
        "Refuserer å kjøre tester mot ikke-test database. "
        "Sett TEST_DATABASE_URL til en *_test-database."
    )
os.environ.setdefault("APP_SECRET_KEY", "test-secret")
os.environ.setdefault("APP_ENV", "development")
# Testene skal ikke avhenge av Redis/Celery-kø.
os.environ["USE_CELERY_BACKGROUND"] = "false"

# Importeres etter at env er satt.
from app.core.config import get_settings
from app.core.database import Base, get_session
from app.main import create_app


@pytest.fixture(scope="session")
def event_loop_policy():
    """Bruk default policy — pytest-asyncio håndterer loop per test."""
    import asyncio

    return asyncio.get_event_loop_policy()


@pytest_asyncio.fixture
async def db_engine():
    """Frisk database per test — drop alle tabeller, opprett dem på nytt."""
    settings = get_settings()
    engine = create_async_engine(settings.database_url, future=True)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)
        await conn.run_sync(Base.metadata.create_all)
    yield engine
    await engine.dispose()


@pytest_asyncio.fixture
async def db_session(db_engine) -> AsyncIterator[AsyncSession]:
    factory = async_sessionmaker(db_engine, expire_on_commit=False)
    async with factory() as session:
        yield session


@pytest_asyncio.fixture
async def app(db_engine, tmp_path: Path) -> AsyncIterator[FastAPI]:
    """FastAPI-app med override av database-session og upload-mappe."""
    factory = async_sessionmaker(db_engine, expire_on_commit=False)

    async def _override_session() -> AsyncIterator[AsyncSession]:
        async with factory() as session:
            try:
                yield session
            except Exception:
                await session.rollback()
                raise

    # Bruk en temporær upload-mappe per test slik at vi ikke skitner til disken.
    upload_dir = tmp_path / "uploads"
    upload_dir.mkdir()
    get_settings().upload_dir = upload_dir

    test_app = create_app()
    test_app.dependency_overrides[get_session] = _override_session
    yield test_app
    test_app.dependency_overrides.clear()


@pytest_asyncio.fixture
async def client(app: FastAPI) -> AsyncIterator[AsyncClient]:
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        yield ac


@pytest.fixture
def fixtures_dir() -> Path:
    """Sti til mappen med test-PDF-er og andre statiske ressurser."""
    return Path(__file__).parent / "fixtures"


@pytest.fixture
def sample_pdf_bytes() -> bytes:
    """Minimal gyldig PDF-byte-streng for opplastings-tester.

    Dette er en av de minste gyldige PDF-strukturene som finnes — tilstrekkelig
    for at pdfplumber kan åpne den. For OCR-tester bruker vi heller ekte
    skannede PDF-er fra `tests/fixtures/`.
    """
    return (
        b"%PDF-1.4\n"
        b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
        b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
        b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]"
        b"/Contents 4 0 R/Resources<<>>>>endobj\n"
        b"4 0 obj<</Length 44>>stream\n"
        b"BT /F1 12 Tf 100 700 Td (Hello Invoice) Tj ET\n"
        b"endstream\nendobj\n"
        b"xref\n0 5\n"
        b"0000000000 65535 f \n"
        b"0000000009 00000 n \n"
        b"0000000052 00000 n \n"
        b"0000000101 00000 n \n"
        b"0000000183 00000 n \n"
        b"trailer<</Size 5/Root 1 0 R>>\n"
        b"startxref\n270\n%%EOF"
    )


@pytest.fixture
def temp_pdf(tmp_path: Path, sample_pdf_bytes: bytes) -> Path:
    pdf_path = tmp_path / "sample.pdf"
    pdf_path.write_bytes(sample_pdf_bytes)
    return pdf_path


@pytest.fixture
def sample_xlsx_bytes() -> bytes:
    """Minimal gyldig XLSX-byte-streng for opplastings- og parser-tester.

    Bruker openpyxl for å generere en ekte ZIP/OOXML-struktur slik at
    både magic-byte-sjekk (PK\\x03\\x04) og Docling-konvertering fungerer.
    """
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws["A1"] = "Invoice Number"
    ws["B1"] = "Amount"
    ws["C1"] = "Currency"
    ws["A2"] = "INV-001"
    ws["B2"] = 1000.00
    ws["C2"] = "NOK"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def temp_xlsx(tmp_path: Path, sample_xlsx_bytes: bytes) -> Path:
    xlsx_path = tmp_path / "sample.xlsx"
    xlsx_path.write_bytes(sample_xlsx_bytes)
    return xlsx_path
